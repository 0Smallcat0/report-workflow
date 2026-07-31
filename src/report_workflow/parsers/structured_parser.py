"""Structured parser for CSV, XLSX, JSON files."""
import csv
import json
import tomllib

# What people type in a cell they have not filled in yet. Deliberately does not
# include "0" or "false" — those are results, not absences — nor "?" , which a
# survey may well use as a real answer code.
PLACEHOLDER_CELL_VALUES = frozenset({
    "", "-", "--", "---", "—", "–", "n/a", "n.a.", "na", "nil",
    "null", "none", "nan", "tbd", "待填", "無", "无", "未測", "未测",
})


def is_placeholder_value(value: object) -> bool:
    """True when a cell holds a stand-in rather than a value."""
    return str(value).strip().lower() in PLACEHOLDER_CELL_VALUES


def _disambiguate_headers(headers: list[str]) -> list[str]:
    """Make repeated column names distinct, keeping the first as written.

    An instrument export names every thermocouple "Temperature (°C)".
    DictReader keys rows by name, so the second and third readings overwrote
    the first and left the row holding one temperature out of three — data
    the source contained, destroyed at ingestion, with the table then
    reported as having four columns instead of six. pandas already renames
    duplicates for .xlsx, so only the CSV path lost them.
    """
    seen: dict[str, int] = {}
    out: list[str] = []
    for header in headers:
        name = str(header)
        seen[name] = seen.get(name, 0) + 1
        out.append(name if seen[name] == 1 else f"{name} [{seen[name]}]")
    return out


def _cell_count(row) -> int:
    filled = 0
    for cell in row:
        text = str(cell).strip()
        if text and text.lower() != "nan":
            filled += 1
    return filled


def _leading_noise_rows(rows: list) -> int:
    """How many rows sit above a table's real header.

    A monthly export opens with its own name in A1 and a blank line under it.
    Read as the header, that title became the first column's name, the other
    columns became Unnamed, and the real header — 月份, 不良率(%), 產量(件) —
    was demoted to a citable data row, taking the unit out of the column name
    that the unit checks read. In CSV it was worse: a one-cell header truncated
    every data row to one column and the other two measurements were gone.

    A title fills one cell and a spacer none, while any header worth the name
    spans at least two, so only leading rows holding at most one value are
    skipped, and only when the row directly below them is the widest in the
    table. A header with a blank corner cell still holds two or more values and
    is never passed over.
    """
    widest = max((_cell_count(row) for row in rows), default=0)
    if widest < 2:
        return 0
    skipped = 0
    for row in rows[:5]:
        if _cell_count(row) <= 1 and skipped + 1 < len(rows):
            skipped += 1
            continue
        break
    if not skipped or _cell_count(rows[skipped]) != widest:
        return 0
    return skipped


def parse_csv(file_path: str) -> list[dict]:
    """Parse CSV file with the standard library."""
    import io

    from .source_text import read_source_text

    rows = list(csv.reader(io.StringIO(read_source_text(file_path), newline="")))
    if not rows:
        return []
    rows = rows[_leading_noise_rows(rows):]
    headers = _disambiguate_headers(rows[0])
    return [
        dict(zip(headers, row))
        for row in rows[1:]
        if any(str(cell).strip() for cell in row)
    ]


def _merged_cell_values(file_path: str) -> dict[str, dict[tuple[int, int], object]]:
    """What each merged range shows, per sheet, keyed by zero-based cell.

    A merge reaches pandas as the top-left value and blanks beside it, so the
    group column of a course data sheet — 組別 A spanning its three runs, B
    spanning its three — arrived as two readings with a group and four with
    NaN. What the sheet plainly shows to anyone looking at it was destroyed at
    ingestion, and the rows that survived carried no way to say whose they were.

    Carrying values down into every gap would be a different and worse answer:
    a run nobody measured would inherit the reading before it. Only the cells a
    merge actually covers are filled, which is why the ranges are read out of
    the file rather than guessed from where the blanks fall.
    """
    from openpyxl import load_workbook

    filled: dict[str, dict[tuple[int, int], object]] = {}
    try:
        book = load_workbook(file_path, data_only=True)
    except Exception:
        # A workbook openpyxl declines may still be readable by pandas, and a
        # sheet with no merges loses nothing by this coming back empty.
        return filled
    try:
        for sheet in book.worksheets:
            cells: dict[tuple[int, int], object] = {}
            for merged in sheet.merged_cells.ranges:
                value = sheet.cell(merged.min_row, merged.min_col).value
                if value is None:
                    continue
                top = merged.min_row - 1
                for row in range(merged.min_row, merged.max_row + 1):
                    for col in range(merged.min_col, merged.max_col + 1):
                        cells[(row - 1, col - 1)] = (value, top)
            if cells:
                filled[str(sheet.title)] = cells
    finally:
        book.close()
    return filled


def parse_xlsx(file_path: str) -> list[dict]:
    """Parse XLSX file using pandas — every sheet, not only the first.

    A workbook that keeps one year per tab had its other tabs read by nobody:
    those rows never reached the ledger and nothing said they existed. When
    more than one sheet holds data each record carries the name of the sheet it
    came from, because rows reading 1.8 and 4.2 with no way to tell which year
    is the confusion that gets the wrong one cited. A single-sheet workbook —
    the ordinary case — is parsed exactly as before, with no added key.

    Merged cells are filled in after the header row has been located, not
    before: a report's title merged across the width of its table shows one
    value, and filling it first would make that row as wide as the table and so
    indistinguishable from the header sitting under it.

    The fill is applied to the frame pandas has already typed rather than to the
    raw cell grid. Reading the grid untyped loses the column's own dtype, and
    3.0 measured alongside 2.5 came back as 3 — the same reading, reported to
    one fewer significant figure.
    """
    import pandas as pd

    merged = _merged_cell_values(file_path)
    book = pd.read_excel(file_path, sheet_name=None, header=None)
    sheets: list[tuple[str, list[dict]]] = []
    for name, raw in book.items():
        if raw.empty:
            continue
        skip = _leading_noise_rows(raw.values.tolist())
        frame = pd.read_excel(file_path, sheet_name=name, header=skip)
        cells = merged.get(str(name), {})
        if cells:
            # A merge over the header row names the columns it covers; the same
            # label then appears twice, which is what the CSV and DOCX readers
            # already number rather than let one column overwrite the other.
            columns = [str(column) for column in frame.columns]
            for (row, col), (value, _top) in cells.items():
                if row == skip and 0 <= col < len(columns):
                    columns[col] = str(value)
            frame.columns = _disambiguate_headers(columns)
            for (row, col), (value, top) in cells.items():
                # A merge that reaches up into the header row is the header —
                # a two-row header writes 有效度(%) down over the row beneath it.
                # Filling that as data put the column's own name into the first
                # reading, and pandas refused a string in a float column, so the
                # whole sheet came back unreadable: a lab sheet with a stacked
                # header produced no evidence at all. Only merges that lie
                # entirely below the header describe readings.
                if top <= skip:
                    continue
                data_row = row - skip - 1
                if 0 <= data_row < len(frame) and 0 <= col < len(frame.columns):
                    frame.iloc[data_row, col] = value
        records = frame.to_dict(orient="records")
        if records:
            sheets.append((str(name), records))

    if len(sheets) <= 1:
        return sheets[0][1] if sheets else []

    combined: list[dict] = []
    for name, records in sheets:
        for record in records:
            key = "sheet"
            suffix = 2
            while key in record:
                key = f"sheet [{suffix}]"
                suffix += 1
            combined.append({key: name, **record})
    return combined


def parse_json(file_path: str) -> list[dict]:
    """Parse JSON file."""
    from .source_text import read_source_text

    data = json.loads(read_source_text(file_path))
    if isinstance(data, list):
        return data
    elif isinstance(data, dict):
        return [data]
    return []


def parse_toml(file_path: str) -> list[dict]:
    """Parse TOML files such as pyproject.toml."""
    with open(file_path, "rb") as f:
        data = tomllib.load(f)
    if isinstance(data, dict):
        return [data]
    return []


def _json_safe(record: dict) -> dict:
    """A cell nobody filled in, written as null rather than as NaN.

    ``json.dumps`` emits a bare ``NaN`` by default, which is not JSON: no
    strict parser accepts it. The evidence ledger is a delivered artifact meant
    to be read by whoever checks the report, so a run that was never measured
    put a line in it that a reader in any other language would refuse outright.
    ``null`` says the same thing and says it legally — and it renders as an
    empty cell instead of the word "nan".
    """
    return {
        key: None if isinstance(value, float) and value != value else value
        for key, value in record.items()
    }


def parse_structured(file_path: str, file_type: str) -> dict:
    """Parse structured file and return content blocks."""
    try:
        if file_type == "csv":
            records = parse_csv(file_path)
        elif file_type == "xlsx":
            records = parse_xlsx(file_path)
        elif file_type == "json":
            records = parse_json(file_path)
        elif file_type == "toml":
            records = parse_toml(file_path)
        else:
            return {"blocks": [], "error": f"Unsupported file type: {file_type}"}

        blocks = []
        content_lines = []
        for i, record in enumerate(records):
            record = _json_safe(record)
            line = json.dumps(record, ensure_ascii=False)
            headers = [str(key) for key in record.keys()]
            values = ["" if value is None else str(value) for value in record.values()]
            # A row exported before anyone filled it in carries column names and
            # nothing else. It used to become an evidence entry all the same, so
            # an untouched export could clear the "at least 5 evidence entries"
            # bar with five rows of dashes. Note "0" is a measurement, not a
            # blank — only the placeholders below count as absent.
            if values and all(is_placeholder_value(value) for value in values):
                continue
            content_lines.append(line)
            blocks.append({
                "block_id": f"block_{i}",
                "block_type": "table_row",
                "content": line,
                "page_number": None,
                "table_data": [headers, values] if headers or values else None
            })

        return {
            "blocks": blocks,
            "raw_content": "\n".join(content_lines),
            "success": True
        }
    except Exception as e:
        return {"blocks": [], "error": str(e), "success": False}
